import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta

# Crear workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ==================== HOJA 1: CALENDARIO ANDREA ====================
ws_andrea = wb.create_sheet("Calendario Andrea", 0)

# Configurar columnas
ws_andrea.column_dimensions['A'].width = 12
for col in range(2, 7):
    ws_andrea.column_dimensions[chr(64+col)].width = 18

# Título
ws_andrea['A1'] = "CALENDARIO ANDREA - ABRIL 2026"
ws_andrea['A1'].font = Font(bold=True, size=12)
ws_andrea.merge_cells('A1:F1')

# Dias de la semana (lunes a viernes)
today = datetime(2026, 4, 6)  # Lunes 6 de abril
ws_andrea['A3'] = "HORA"
for col_num, day_offset in enumerate(range(5), 2):
    day = today + timedelta(days=day_offset)
    col_letter = chr(64 + col_num)
    ws_andrea[f'{col_letter}3'] = day.strftime("%a %d/%m")
    ws_andrea[f'{col_letter}3'].font = Font(bold=True)
    ws_andrea[f'{col_letter}3'].alignment = Alignment(horizontal='center')

# Horarios (8 a 20 hs)
horarios = ["8:00", "9:00", "10:00", "11:00", "12:00", "13:00", "14:30", "15:30", "16:30", "17:30", "18:30", "19:30", "20:30"]
for row_num, hora in enumerate(horarios, 4):
    ws_andrea[f'A{row_num}'] = hora
    ws_andrea[f'A{row_num}'].font = Font(bold=True)

# Llenar con algunos pacientes de ejemplo
pacientes_andrea = {
    (4, 2): "Federico",      # Lunes 8hs
    (5, 2): "Ismael",        # Martes 8hs
    (10, 2): "Morena",       # Viernes 16:30
    (11, 3): "Vicky",        # Sábado 17:30
    (12, 4): "Sofía",        # Domingo 18:30
}

# Colores
libre_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde
ocupado_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Celeste
no_disp_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo

# Rellena celdas
for row in range(4, 4 + len(horarios)):
    for col in range(2, 7):
        cell = ws_andrea.cell(row=row, column=col)
        col_letter = chr(64 + col)
        cell_key = (row - 3, col - 1)

        if cell_key in pacientes_andrea:
            cell.value = pacientes_andrea[cell_key]
            cell.fill = ocupado_fill
        else:
            cell.value = "[LIBRE]"
            cell.fill = libre_fill

        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ==================== HOJA 2: CALENDARIO YAMILA ====================
ws_yamila = wb.create_sheet("Calendario Yamila", 1)

ws_yamila.column_dimensions['A'].width = 12
for col in range(2, 7):
    ws_yamila.column_dimensions[chr(64+col)].width = 18

ws_yamila['A1'] = "CALENDARIO YAMILA - ABRIL 2026"
ws_yamila['A1'].font = Font(bold=True, size=12)
ws_yamila.merge_cells('A1:F1')

ws_yamila['A3'] = "HORA"
for col_num, day_offset in enumerate(range(5), 2):
    day = today + timedelta(days=day_offset)
    col_letter = chr(64 + col_num)
    ws_yamila[f'{col_letter}3'] = day.strftime("%a %d/%m")
    ws_yamila[f'{col_letter}3'].font = Font(bold=True)
    ws_yamila[f'{col_letter}3'].alignment = Alignment(horizontal='center')

for row_num, hora in enumerate(horarios, 4):
    ws_yamila[f'A{row_num}'] = hora
    ws_yamila[f'A{row_num}'].font = Font(bold=True)

# Llenar con pacientes de Yamila
pacientes_yamila = {
    (10, 4): "Nicolas Cattaneo",  # Miercoles 16:30
}

for row in range(4, 4 + len(horarios)):
    for col in range(2, 7):
        cell = ws_yamila.cell(row=row, column=col)
        cell_key = (row - 3, col - 1)

        if cell_key in pacientes_yamila:
            cell.value = pacientes_yamila[cell_key]
            cell.fill = ocupado_fill
        else:
            cell.value = "[LIBRE]"
            cell.fill = libre_fill

        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ==================== HOJA 3: PACIENTES DERIVADOS ====================
ws_pacientes = wb.create_sheet("Pacientes Derivados", 2)

headers = ["Semana", "Fecha Der.", "Paciente", "Tipo", "Profesional", "Telefono", "Consultorio", "Hora", "Asistio", "Cancelado", "Monto", "Estado Pago"]
for col_num, header in enumerate(headers, 1):
    cell = ws_pacientes.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

ws_pacientes.column_dimensions['A'].width = 12
ws_pacientes.column_dimensions['B'].width = 12
ws_pacientes.column_dimensions['C'].width = 18
ws_pacientes.column_dimensions['D'].width = 12
ws_pacientes.column_dimensions['E'].width = 15
ws_pacientes.column_dimensions['F'].width = 12
ws_pacientes.column_dimensions['G'].width = 12
ws_pacientes.column_dimensions['H'].width = 8
ws_pacientes.column_dimensions['I'].width = 10
ws_pacientes.column_dimensions['J'].width = 12
ws_pacientes.column_dimensions['K'].width = 10
ws_pacientes.column_dimensions['L'].width = 14

# Ejemplo de fila
row_data = [
    "3-7 Abr",
    "3 Abr",
    "Federico",
    "Sesion",
    "Andrea",
    "1123456",
    "C2",
    "8:00",
    "SI",
    "NO",
    "$15k",
    "Cobrado"
]

for col_num, value in enumerate(row_data, 1):
    cell = ws_pacientes.cell(row=2, column=col_num)
    cell.value = value
    cell.alignment = Alignment(horizontal='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ==================== HOJA 4: RESUMEN MENSUAL ====================
ws_resumen = wb.create_sheet("Resumen Mensual", 3)

ws_resumen['A1'] = "PSCIELO - RESUMEN ABRIL 2026"
ws_resumen['A1'].font = Font(bold=True, size=14)
ws_resumen.merge_cells('A1:B1')

ws_resumen['A3'] = "INGRESOS"
ws_resumen['A3'].font = Font(bold=True, size=11)

ws_resumen['A4'] = "Alquileres"
ws_resumen['B4'] = "$850,000"

ws_resumen['A5'] = "Comisiones Andrea"
ws_resumen['B5'] = "$245,000"

ws_resumen['A6'] = "Comisiones Yamila"
ws_resumen['B6'] = "$178,000"

ws_resumen['A7'] = "TOTAL INGRESOS"
ws_resumen['A7'].font = Font(bold=True)
ws_resumen['B7'].font = Font(bold=True)
ws_resumen['B7'] = "$1,273,000"

ws_resumen['A9'] = "GASTOS"
ws_resumen['A9'].font = Font(bold=True, size=11)

ws_resumen['A10'] = "Alquiler local"
ws_resumen['B10'] = "$2,269,000"

ws_resumen['A11'] = "Electricidad (EPEC)"
ws_resumen['B11'] = "$85,000"

ws_resumen['A12'] = "Gas (Ecogas)"
ws_resumen['B12'] = "$5,000"

ws_resumen['A13'] = "TOTAL GASTOS"
ws_resumen['A13'].font = Font(bold=True)
ws_resumen['B13'].font = Font(bold=True)
ws_resumen['B13'] = "$2,359,000"

ws_resumen['A15'] = "RESULTADO"
ws_resumen['A15'].font = Font(bold=True, size=12, color="FF0000")
ws_resumen['B15'] = "-$1,086,000"
ws_resumen['B15'].font = Font(bold=True, size=12, color="FF0000")

ws_resumen.column_dimensions['A'].width = 25
ws_resumen.column_dimensions['B'].width = 20

# Guardar
wb.save(r'D:\pscielo\PSCIELO_Template_Gestion.xlsx')
print("Template creado exitosamente!")
